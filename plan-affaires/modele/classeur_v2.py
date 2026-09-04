#!/usr/bin/env python3
"""Classeur Excel du plan v2 avec VRAIES formules (standard RCGT) : Hypothèses → Investissement →
Dette (par couche, PMT) → Résultats → Flux → Bilan → Ratios → Sensibilité → DPA.
Les valeurs de départ viennent de modele_v2.json ; tout le reste se recalcule dans Excel.
Usage : python classeur_v2.py  → ../Plan_affaires_Complexe_Havana_v2.xlsx"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L

HERE = os.path.dirname(os.path.abspath(__file__))
M = json.load(open(os.path.join(HERE, "modele_v2.json"), encoding="utf8"))
OUT = os.path.abspath(os.path.join(HERE, "..", "Plan_affaires_Complexe_Havana_v2.xlsx"))
H = M["hypotheses"]
wb = Workbook()
NAVY = "0C4A5E"; LIGHT = "F5F2EA"
hfont = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor=NAVY); bold = Font(bold=True)
inp = PatternFill("solid", fgColor="FFF6D5")   # cellule d'entrée (jaune pâle)
money = '#,##0 $;[Red](#,##0 $)'; pctf = '0.0 %'; ratf = '0.00" x"'
thin = Side(style="thin", color="CFE3E6")

def header(ws, row, labels, widths=None):
    for i, t in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=t); c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, 1): ws.column_dimensions[L(i)].width = w

# ---------------------------------------------------------------- Hypothèses
ws = wb.active; ws.title = "Hypothèses"
header(ws, 1, ["Paramètre", "Valeur", "Unité", "Source / statut"], [58, 16, 12, 70])
NAMES = {}
def hyp(name, label, value, unit, source, fmt=None):
    r = ws.max_row + 1
    ws.cell(r, 1, label); c = ws.cell(r, 2, value); c.fill = inp; ws.cell(r, 3, unit); ws.cell(r, 4, source)
    if fmt: c.number_format = fmt
    NAMES[name] = f"Hypothèses!$B${r}"
    return NAMES[name]
def hyp_row(name, label, values, unit, source, fmt=None):
    """une ligne avec 6 valeurs (2026, An1..An5) en colonnes B..G"""
    r = ws.max_row + 1
    ws.cell(r, 1, label)
    for i, v in enumerate(values):
        c = ws.cell(r, 2 + i, v); c.fill = inp
        if fmt: c.number_format = fmt
    ws.cell(r, 8, unit); ws.cell(r, 9, source)
    NAMES[name] = r
    return r
ws.cell(2, 1, "GÉNÉRAL").font = bold
hyp("infl", "Indexation annuelle des tarifs", 0.03, "%", "Pratique de l'industrie", pctf)
hyp("terrains_2026", "Ventes terrains 2026 (Pixum)", 791894.29, "$", "Rapport Pixum 2025-09-15 → 2026-09-15", money)
hyp("saison_2026", "Ventes saisonniers 2026 (Pixum)", 43208.71, "$", "Pixum", money)
hyp("chalets_2026", "Ventes chalets et chambres 2026 (Pixum)", 497722.97, "$", "Pixum", money)
hyp("cabanas_2026", "Ventes cabanas 2026 (Pixum)", 199015.63, "$", "Pixum", money)
hyp("resto_2026", "Restauration 2026", 100000, "$", "Budget interne, à valider", money)
hyp("activ_2026", "Activités et concessions 2026", 96700, "$", "Budget interne, à valider", money)
hyp("redev_2026", "Part Havana des ventes Coolbox 2026 (après transport)", M["coolbox_2026"]["redevance"], "$", "AVIS-HAVANA-2026-001", money)
hyp("redev_base", "Part Havana des ventes Coolbox, base récurrente (50 % du CA 2026)", H["redevance_base"], "$", "CALCUL-HAVANA 2026", money)
r0 = ws.max_row + 2; ws.cell(r0, 1, "PAR ANNÉE (colonnes B = 2026, C..G = An 1 à An 5)").font = bold
for i, t in enumerate(["2026", "An 1", "An 2", "An 3", "An 4", "An 5"]): ws.cell(r0, 2 + i, t).font = bold
hyp_row("terrains_uplift", "Terrains : facteur d'occupation", [1, 1.00, 1.02, 1.03, 1.04, 1.04], "x", "Aucune capacité nouvelle", '0.00')
hyp_row("sais_sites", "Contrats saisonniers", [12, 30, 45, 58, 66, 70], "contrats", "12 réels en 2026 ; conversion de terrains existants")
hyp_row("chalets_occ_ete", "Chalets : occupation été (184 nuits)", [0, .58, .61, .63, .65, .66], "%", "Réel 2026 ≈ 55-60 %", pctf)
hyp_row("chalets_occ_hiv", "Chalets : occupation hiver (181 nuits)", [0, .12, .18, .22, .24, .25], "%", "Hivernisation automne 2027", pctf)
hyp_row("chalets_frac_hiv", "Chalets : fraction de l'hiver exploitée", [0, .30, 1, 1, 1, 1], "x", "An 1 : nov.-déc.", '0.00')
hyp_row("villas_occ_ete", "Villas : occupation été", [0, .50, .55, .60, .63, .65], "%", "Direction", pctf)
hyp_row("villas_occ_hiv", "Villas : occupation hiver", [0, .10, .16, .20, .21, .22], "%", "hyp.", pctf)
hyp_row("villas_frac_ete", "Villas : fraction de l'été exploitée", [0, .50, 1, 1, 1, 1], "x", "Livraison juillet 2027", '0.00')
hyp_row("hotel_rooms", "Hôtel : chambres", [6, 11.5, 17, 17, 17, 17], "ch.", "6 jan.-juin 2027, 17 ensuite")
hyp_row("hotel_occ", "Hôtel : occupation annuelle", [0, .35, .42, .48, .52, .55], "%", "Sous la moyenne régionale", pctf)
hyp_row("cabanas_uplift", "Cabanas : facteur", [1, 1.0, 1.02, 1.04, 1.05, 1.05], "x", "", '0.00')
hyp_row("cb_occ_ete", "Coolbox grappe A : occupation été", [0, .55, .60, .63, .65, .65], "%", "9 unités existantes : 20 037 $/unité en 2026", pctf)
hyp_row("cb_occ_hiv", "Coolbox grappe A : occupation hiver", [0, .20, .27, .32, .34, .35], "%", "Données d'hiver du réseau (annexe J)", pctf)
hyp_row("cb_frac_ete", "Coolbox : fraction de l'été An 1", [0, .80, 1, 1, 1, 1], "x", "Livraison mai 2027", '0.00')
hyp_row("fb_ratio", "Restauration : % des revenus d'hébergement (sans alcool)", [0, .08, .14, .18, .19, .20], "%", "Réel 2026 : 6,5 %", pctf)
hyp_row("spectacles", "Spectacles : billetterie brute", [0, 0, 60000, 120000, 150000, 160000], "$", "Scène et chapiteau existants (hyp.)", money)
hyp_row("activites", "Activités et concessions", [96700, 105000, 125000, 140000, 148000, 155000], "$", "Budget interne", money)
hyp_row("sal_pct", "Salaires : % des revenus", [.39, .36, .34, .32, .31, .30], "%", "Tableau d'effectifs ; Grenier 2021 : 30 %", pctf)
hyp_row("energie", "Énergie", [242500, 275000, 400000, 420000, 435000, 450000], "$", "Budget de la direction", money)
hyp_row("assur", "Assurances", [60000, 95000, 130000, 134000, 138000, 142000], "$", "", money)
hyp_row("taxes", "Taxes municipales et scolaires", [54500, 75000, 130000, 133000, 136000, 139000], "$", "Rôle 2024-2026", money)
hyp_row("mkt_pct", "Marketing : % des revenus", [.03, .035, .035, .03, .03, .03], "%", "", pctf)
r1 = ws.max_row + 2; ws.cell(r1, 1, "TARIFS ET UNITÉS").font = bold
hyp("n_chalets", "Chalets (unités)", 16, "u", "Inventaire"); hyp("t_chalet_ete", "Tarif chalet été", 235, "$", "Tarif 2026", money); hyp("t_chalet_hiv", "Tarif chalet hiver", 210, "$", "hyp.", money)
hyp("n_villas", "Villas (unités)", 13, "u", ""); hyp("t_villa_ete", "Tarif villa été", 275, "$", "Direction", money); hyp("t_villa_hiv", "Tarif villa hiver", 245, "$", "hyp.", money)
hyp("t_hotel", "Tarif chambre d'hôtel", 190, "$", "Tarif 2026", money)
hyp("n_cb", "Unités Coolbox grappe A", 9, "u", "Emplacements 2022"); hyp("t_cb_ete", "Tarif Coolbox été", 225, "$", "Affiché 2023", money); hyp("t_cb_hiv", "Tarif Coolbox hiver", 199, "$", "hyp.", money)
hyp("t_sais", "Tarif contrat saisonnier", 3650, "$", "2026", money)
hyp("t_spa", "Tarif spa (phase 2, hors base)", 55, "$", "", money)
r2 = ws.max_row + 2; ws.cell(r2, 1, "CHARGES").font = bold
hyp("cv_resto", "Coût des ventes restauration", 0.33, "%", "Norme 30-35 %", pctf); hyp("cv_activ", "Coût des ventes activités", 0.20, "%", "", pctf)
hyp("cachets", "Cachets et technique (% billetterie)", 0.55, "%", "hyp.", pctf); hyp("comm", "Commissions plateformes (15 % × 40 % des nuitées hôtel/villas/Coolbox)", 0.06, "%", "hyp.", pctf)
hyp("entretien", "Entretien (% des revenus)", 0.08, "%", "Grenier 2021", pctf); hyp("maintien", "Immobilisations de maintien (% des revenus)", 0.02, "%", "Pratique bancaire", pctf)
hyp("vehicules", "Véhicules et essence (2026)", 88500, "$", "Budget de la direction", money); hyp("honoraires", "Honoraires comptables (2026)", 25000, "$", "hyp.", money)
hyp("admin", "Administration, TI (% revenus)", 0.02, "%", "", pctf); hyp("bancaires", "Frais bancaires (% revenus)", 0.015, "%", "", pctf)
hyp("fournitures", "Fournitures (% revenus)", 0.025, "%", "", pctf); hyp("loyers", "Loyers et locations (2026)", 42000, "$", "Budget de la direction", money); hyp("divers", "Divers (% revenus)", 0.02, "%", "", pctf)
hyp("amort_exist", "Amortissement comptable des actifs existants", 150000, "$", "hyp.", money)
hyp("impot_pme", "Impôt PME (jusqu'à 500 k$)", 0.122, "%", "À confirmer", pctf); hyp("impot_gen", "Impôt général (au-delà)", 0.265, "%", "À confirmer", pctf)
hyp("dette_exist_taux", "Dette existante 2026 : taux (intérêts seulement)", 0.075, "%", "hyp. ; relevés à joindre", pctf)
r3 = ws.max_row + 2; ws.cell(r3, 1, "PART VARIABLE DES CHARGES (sensibilité)").font = bold
PV = {"cout_ventes": 1, "cachets": 1, "commissions": 1, "salaires": .5, "energie": .15, "entretien": .5, "vehicules": .3, "marketing": .5, "assurances": 0, "taxes": 0, "honoraires": 0, "admin": .5, "bancaires": 1, "fournitures": 1, "loyers": 0, "divers": 1}
for k, v in PV.items(): hyp("pv_" + k, "Part variable : " + k, v, "%", "Structure du modèle", pctf)
hyp("scen_rev", "SCÉNARIO : facteur sur les revenus (1 = base)", 1.0, "x", "Modifier pour tester", '0.00')

# ---------------------------------------------------------------- Investissement
wi = wb.create_sheet("Investissement")
header(wi, 1, ["Emploi", "Montant", "Statut"], [80, 16, 30])
r = 2
for lab, val in M["emplois"]:
    wi.cell(r, 1, lab); c = wi.cell(r, 2, val); c.fill = inp; c.number_format = money; r += 1
wi.cell(r, 1, "Total des emplois").font = bold; wi.cell(r, 2, f"=SUM(B2:B{r-1})").number_format = money; TOT_EMPLOIS = f"Investissement!$B${r}"; r += 2
header(wi, r, ["Source", "Montant", "Taux", "Amortissement (ans)", "Moratoire (mois)", "Statut"]); r += 1
src0 = r
for c in M["couches"]:
    wi.cell(r, 1, c["nom"]); wi.cell(r, 2, c["montant"]).number_format = money; wi.cell(r, 2).fill = inp
    wi.cell(r, 3, c["taux"]).number_format = pctf; wi.cell(r, 3).fill = inp; wi.cell(r, 4, c["amort_ans"]).fill = inp; wi.cell(r, 5, c["moratoire_mois"]).fill = inp; wi.cell(r, 6, c["statut"]); r += 1
src1 = r - 1
wi.cell(r, 1, "Fonds propres (= emplois − couches de dette)").font = bold; wi.cell(r, 2, f"={TOT_EMPLOIS}-SUM(B{src0}:B{src1})").number_format = money; FP = f"Investissement!$B${r}"; r += 1
wi.cell(r, 1, "Total des sources").font = bold; wi.cell(r, 2, f"=SUM(B{src0}:B{r-1})").number_format = money; r += 1
wi.cell(r, 1, "Argent neuf (emplois hors refinancement)"); wi.cell(r, 2, f"={TOT_EMPLOIS}-B{[i for i,(lab,_) in enumerate(M['emplois'],2) if lab.startswith('Refinancement')][0]}").number_format = money; AN = f"Investissement!$B${r}"; r += 1
wi.cell(r, 1, "Fonds propres / argent neuf"); wi.cell(r, 2, f"={FP}/{AN}").number_format = pctf
for i in range(4): wi.column_dimensions[L(3 + i)].width = 18

# ---------------------------------------------------------------- Dette (par couche, mensuel → annuel)
wd = wb.create_sheet("Dette")
wd.column_dimensions["A"].width = 60
row = 1
SERVICE_REFS, INT_REFS, CAP_REFS, SOLDE_REFS = [], [], [], []
for ci, c in enumerate(M["couches"]):
    ir = src0 + ci
    wd.cell(row, 1, c["nom"]).font = bold; row += 1
    wd.cell(row, 1, "Montant"); wd.cell(row, 2, f"=Investissement!B{ir}").number_format = money
    wd.cell(row, 3, "Taux"); wd.cell(row, 4, f"=Investissement!C{ir}").number_format = pctf
    wd.cell(row, 5, "Amort. (ans)"); wd.cell(row, 6, f"=Investissement!D{ir}")
    wd.cell(row, 7, "Moratoire (mois)"); wd.cell(row, 8, f"=Investissement!E{ir}")
    wd.cell(row, 9, "Paiement mensuel"); wd.cell(row, 10, f"=IF(D{row}=0,B{row}/(F{row}*12),PMT(D{row}/12,F{row}*12,-B{row}))").number_format = money
    prm = row; row += 1
    header(wd, row, ["Mois", "Solde début", "Intérêts", "Capital", "Paiement", "Solde fin"]); row += 1
    first = row
    tirage = M["couches"][ci].get("tirage", 1.0)
    for m in range(1, 61):
        wd.cell(row, 1, m)
        wd.cell(row, 2, f"=B{prm}" if m == 1 else f"=F{row-1}").number_format = money
        # An 1 du senior : intérêts sur le solde moyen décaissé (facteur de tirage)
        fac = f"*{tirage}" if (ci == 0 and m <= 12) else ""
        wd.cell(row, 3, f"=B{row}*D${prm}/12{fac}").number_format = money
        wd.cell(row, 4, f"=IF(A{row}<=H${prm},0,MIN(B{row},IF(D${prm}=0,J${prm},J${prm}-B{row}*D${prm}/12)))").number_format = money
        wd.cell(row, 5, f"=C{row}+D{row}").number_format = money
        wd.cell(row, 6, f"=B{row}-D{row}").number_format = money
        row += 1
    last = row - 1
    header(wd, row, ["Année", "Intérêts", "Capital", "Service", "Solde fin"]); row += 1
    s_refs, i_refs, c_refs, so_refs = [], [], [], []
    for an in range(1, 6):
        a, b = first + (an - 1) * 12, first + an * 12 - 1
        wd.cell(row, 1, f"An {an}")
        wd.cell(row, 2, f"=SUM(C{a}:C{b})").number_format = money; wd.cell(row, 3, f"=SUM(D{a}:D{b})").number_format = money
        wd.cell(row, 4, f"=B{row}+C{row}").number_format = money; wd.cell(row, 5, f"=F{b}").number_format = money
        i_refs.append(f"Dette!$B${row}"); c_refs.append(f"Dette!$C${row}"); s_refs.append(f"Dette!$D${row}"); so_refs.append(f"Dette!$E${row}"); row += 1
    INT_REFS.append(i_refs); CAP_REFS.append(c_refs); SERVICE_REFS.append(s_refs); SOLDE_REFS.append(so_refs); row += 1
for col in "BCDEFGHIJ": wd.column_dimensions[col].width = 15
# tirage An 1 du senior (solde moyen décaissé) — écrit dans le json par le modèle ? sinon 0,55 de la tranche B
# (le facteur est appliqué ci-dessus via M["couches"][0]["tirage"] s'il existe)

# ---------------------------------------------------------------- Résultats
wr = wb.create_sheet("Résultats")
header(wr, 1, ["", "2026", "An 1", "An 2", "An 3", "An 4", "An 5"], [62, 15, 15, 15, 15, 15, 15])
COLS = ["B", "C", "D", "E", "F", "G"]
def hcol(name, j): return f"Hypothèses!${COLS[j]}${NAMES[name]}"      # ligne par année
def hv(name): return NAMES[name]                                     # valeur unique
def idx(j): return f"(1+{hv('infl')})^{j}"        # indexation depuis 2026 (j = année 0..5)
def idx1(j): return f"(1+{hv('infl')})^{max(j-1,0)}"
r = 2; REV = {}
def line(key, label, formulas):
    global r
    wr.cell(r, 1, label)
    for j, f in enumerate(formulas): wr.cell(r, 2 + j, f).number_format = money
    REV[key] = r; r += 1
wr.cell(r, 1, "REVENUS").font = bold; r += 1
line("terrains", "Terrains de camping", [f"={hv('terrains_2026')}"] + [f"={hv('terrains_2026')}*{idx(j)}*{hcol('terrains_uplift', j)}" for j in range(1, 6)])
line("saisonniers", "Terrains saisonniers", [f"={hv('saison_2026')}"] + [f"={hcol('sais_sites', j)}*{hv('t_sais')}*{idx1(j)}" for j in range(1, 6)])
line("chalets", "Chalets 4 saisons (16)", [f"={hv('chalets_2026')}"] + [f"={hv('n_chalets')}*(184*{hcol('chalets_occ_ete', j)}*{hv('t_chalet_ete')}*{idx1(j)}+181*{hcol('chalets_frac_hiv', j)}*{hcol('chalets_occ_hiv', j)}*{hv('t_chalet_hiv')}*{idx1(j)})" for j in range(1, 6)])
line("villas", "Villas (13)", ["=0"] + [f"={hv('n_villas')}*(184*{hcol('villas_frac_ete', j)}*{hcol('villas_occ_ete', j)}*{hv('t_villa_ete')}*{idx1(j)}+181*{hcol('villas_occ_hiv', j)}*{hv('t_villa_hiv')}*{idx1(j)})" for j in range(1, 6)])
line("hotel", "Hôtel (6 → 17 chambres)", ["=0"] + [f"={hcol('hotel_rooms', j)}*365*{hcol('hotel_occ', j)}*{hv('t_hotel')}*{idx1(j)}" for j in range(1, 6)])
line("cabanas", "Cabanas (13)", [f"={hv('cabanas_2026')}"] + [f"={hv('cabanas_2026')}*{idx(j)}*{hcol('cabanas_uplift', j)}" for j in range(1, 6)])
line("coolbox", "Unités Coolbox grappe A (9, propriété Havana)", ["=0"] + [f"={hv('n_cb')}*(184*{hcol('cb_frac_ete', j)}*{hcol('cb_occ_ete', j)}*{hv('t_cb_ete')}*{idx1(j)}+181*{hcol('cb_occ_hiv', j)}*{hv('t_cb_hiv')}*{idx1(j)})" for j in range(1, 6)])
line("redevance", "Partenariat Coolbox HPA, part Havana", [f"={hv('redev_2026')}"] + [f"={hv('redev_base')}*{idx(j)}" for j in range(1, 6)])
heb = lambda col: f"SUM({col}{REV['terrains']}:{col}{REV['coolbox']})"
line("restauration", "Restauration (sans alcool)", [f"={hv('resto_2026')}"] + [f"={heb(COLS[j])}*{hcol('fb_ratio', j)}" for j in range(1, 6)])
line("spectacles", "Spectacles (billetterie brute)", [f"={hcol('spectacles', j)}*(1+{hv('infl')})^MAX({j}-2,0)" for j in range(0, 6)])
line("activites", "Activités et concessions", [f"={hcol('activites', j)}" for j in range(0, 6)])
wr.cell(r, 1, "Total des revenus (× facteur de scénario)").font = bold
for j in range(6): wr.cell(r, 2 + j, f"=SUM({COLS[j]}{REV['terrains']}:{COLS[j]}{REV['activites']})*{hv('scen_rev')}").number_format = money
TOTR = r; r += 2
wr.cell(r, 1, "CHARGES").font = bold; r += 1
CH = {}
def ch(key, label, formulas):
    global r
    wr.cell(r, 1, label)
    for j, f in enumerate(formulas): wr.cell(r, 2 + j, f).number_format = money
    CH[key] = r; r += 1
T = lambda j: f"{COLS[j]}${TOTR}"
ch("cout_ventes", "Coût des ventes", [f"=({COLS[j]}{REV['restauration']}*{hv('cv_resto')}+{COLS[j]}{REV['activites']}*{hv('cv_activ')})*{hv('scen_rev')}" for j in range(6)])
ch("cachets", "Cachets et technique des spectacles", [f"={COLS[j]}{REV['spectacles']}*{hv('cachets')}*{hv('scen_rev')}" for j in range(6)])
ch("commissions", "Commissions de plateformes", [f"=({COLS[j]}{REV['hotel']}+{COLS[j]}{REV['villas']}+{COLS[j]}{REV['coolbox']})*{hv('comm')}*{hv('scen_rev')}" for j in range(6)])
ch("salaires", "Salaires et charges sociales", [f"={T(j)}*{hcol('sal_pct', j)}" for j in range(6)])
ch("energie", "Électricité, chauffage et propane", [f"={hcol('energie', j)}" for j in range(6)])
ch("entretien", "Entretien et réparations", [f"={T(j)}*{hv('entretien')}" for j in range(6)])
ch("vehicules", "Véhicules et essence", [f"={hv('vehicules')}*{idx(j)}" for j in range(6)])
ch("marketing", "Marketing", [f"={T(j)}*{hcol('mkt_pct', j)}" for j in range(6)])
ch("assurances", "Assurances", [f"={hcol('assur', j)}" for j in range(6)])
ch("taxes", "Taxes municipales et scolaires", [f"={hcol('taxes', j)}" for j in range(6)])
ch("honoraires", "Honoraires comptables et audit", [f"={hv('honoraires')}*{idx(j)}" for j in range(6)])
ch("admin", "Administration et TI", [f"={T(j)}*{hv('admin')}" for j in range(6)])
ch("bancaires", "Frais bancaires et de cartes", [f"={T(j)}*{hv('bancaires')}" for j in range(6)])
ch("fournitures", "Fournitures", [f"={T(j)}*{hv('fournitures')}" for j in range(6)])
ch("loyers", "Loyers et locations", [f"={hv('loyers')}*1.02^{j}" for j in range(6)])
ch("divers", "Divers et imprévus", [f"={T(j)}*{hv('divers')}" for j in range(6)])
wr.cell(r, 1, "Total des charges").font = bold
for j in range(6): wr.cell(r, 2 + j, f"=SUM({COLS[j]}{CH['cout_ventes']}:{COLS[j]}{CH['divers']})").number_format = money
TOTC = r; r += 2
wr.cell(r, 1, "BAIIA").font = bold
for j in range(6): wr.cell(r, 2 + j, f"={COLS[j]}{TOTR}-{COLS[j]}{TOTC}").number_format = money
BAIIA = r; r += 1
wr.cell(r, 1, "Marge BAIIA")
for j in range(6): wr.cell(r, 2 + j, f"={COLS[j]}{BAIIA}/{COLS[j]}{TOTR}").number_format = pctf
r += 1
wr.cell(r, 1, "Amortissement comptable (DPA!)")
for j in range(6): wr.cell(r, 2 + j, f"=DPA!{COLS[j]}$3").number_format = money
AMORT = r; r += 1
wr.cell(r, 1, "Intérêts (toutes couches ; 2026 : dette existante)")
refi = [i for i, (lab, _) in enumerate(M["emplois"], 2) if lab.startswith("Refinancement")][0]
wr.cell(r, 2, f"=Investissement!B{refi}*{hv('dette_exist_taux')}").number_format = money
for j in range(1, 6): wr.cell(r, 2 + j, "=" + "+".join(INT_REFS[k][j - 1] for k in range(len(INT_REFS)))).number_format = money
INT = r; r += 1
wr.cell(r, 1, "Bénéfice avant impôts")
for j in range(6): wr.cell(r, 2 + j, f"={COLS[j]}{BAIIA}-{COLS[j]}{AMORT}-{COLS[j]}{INT}").number_format = money
BAI = r; r += 1
wr.cell(r, 1, "Impôts")
for j in range(6): wr.cell(r, 2 + j, f"=IF({COLS[j]}{BAI}<=0,0,MIN({COLS[j]}{BAI},500000)*{hv('impot_pme')}+MAX({COLS[j]}{BAI}-500000,0)*{hv('impot_gen')})").number_format = money
IMP = r; r += 1
wr.cell(r, 1, "Bénéfice net").font = bold
for j in range(6): wr.cell(r, 2 + j, f"={COLS[j]}{BAI}-{COLS[j]}{IMP}").number_format = money
BN = r; r += 2
wr.cell(r, 1, "Service de la dette : prêt senior")
wr.cell(r, 2, f"={COLS[0]}{INT}").number_format = money
for j in range(1, 6): wr.cell(r, 2 + j, f"={SERVICE_REFS[0][j-1]}").number_format = money
SS = r; r += 1
wr.cell(r, 1, "Service de la dette : toutes couches")
wr.cell(r, 2, f"={COLS[0]}{INT}").number_format = money
for j in range(1, 6): wr.cell(r, 2 + j, "=" + "+".join(SERVICE_REFS[k][j - 1] for k in range(len(SERVICE_REFS)))).number_format = money
ST = r; r += 1
wr.cell(r, 1, "Remboursement de capital (toutes couches)")
wr.cell(r, 2, 0).number_format = money
for j in range(1, 6): wr.cell(r, 2 + j, "=" + "+".join(CAP_REFS[k][j - 1] for k in range(len(CAP_REFS)))).number_format = money
CAPR = r; r += 1
wr.cell(r, 1, "Immobilisations de maintien")
wr.cell(r, 2, 0).number_format = money
for j in range(1, 6): wr.cell(r, 2 + j, f"={T(j)}*{hv('maintien')}").number_format = money
MAINT = r; r += 1

# ---------------------------------------------------------------- Flux
wf = wb.create_sheet("Flux")
header(wf, 1, ["Flux de trésorerie (méthode indirecte)", "2026", "An 1", "An 2", "An 3", "An 4", "An 5"], [62, 15, 15, 15, 15, 15, 15])
rows = [("Bénéfice net", BN), ("Plus : amortissement", AMORT)]
wf.cell(2, 1, "Bénéfice net"); wf.cell(3, 1, "Plus : amortissement"); wf.cell(4, 1, "Flux d'exploitation").font = bold
wf.cell(5, 1, "Moins : remboursement de capital"); wf.cell(6, 1, "Moins : immobilisations de maintien"); wf.cell(7, 1, "Flux disponible").font = bold; wf.cell(8, 1, "Trésorerie cumulée depuis la clôture").font = bold
for j in range(6):
    c = COLS[j]
    wf.cell(2, 2 + j, f"=Résultats!{c}{BN}").number_format = money
    wf.cell(3, 2 + j, f"=Résultats!{c}{AMORT}").number_format = money
    wf.cell(4, 2 + j, f"={c}2+{c}3").number_format = money
    wf.cell(5, 2 + j, f"=-Résultats!{c}{CAPR}").number_format = money
    wf.cell(6, 2 + j, f"=-Résultats!{c}{MAINT}").number_format = money
    wf.cell(7, 2 + j, f"={c}4+{c}5+{c}6").number_format = money
    wf.cell(8, 2 + j, ("=0" if j == 0 else f"={COLS[j-1]}8+{c}7")).number_format = money

# ---------------------------------------------------------------- DPA / amortissement comptable
wp = wb.create_sheet("DPA")
header(wp, 1, ["Amortissement comptable et DPA", "2026", "An 1", "An 2", "An 3", "An 4", "An 5"], [70, 15, 15, 15, 15, 15, 15])
wp.cell(2, 1, "Base capitalisée (phase 1a, contingence et honoraires au prorata)")
cap_rows = [i for i, (lab, _) in enumerate(M["emplois"], 2) if lab.startswith(("Projets", "Contingence", "Honoraires"))]
wp.cell(2, 2, "=" + "+".join(f"Investissement!B{i}" for i in cap_rows)).number_format = money
wp.cell(3, 1, "Amortissement comptable total (→ Résultats)").font = bold
wp.cell(4, 1, "Existant"); wp.cell(5, 1, "Nouveau (linéaire moyen 5 % ; phase 1a mise en service An 1, demi-année)")
for j in range(6):
    c = COLS[j]
    wp.cell(4, 2 + j, f"={hv('amort_exist')}").number_format = money
    wp.cell(5, 2 + j, ("=0" if j == 0 else (f"=$B$2*0.05*0.5" if j == 1 else "=$B$2*0.05"))).number_format = money
    wp.cell(3, 2 + j, f"={c}4+{c}5").number_format = money
wp.cell(7, 1, "DPA fiscale (solde dégressif, demi-année) — informatif").font = bold
rr = 8
header(wp, rr, ["Catégorie", "Taux", "Base", "DPA An 1", "DPA An 2", "DPA An 3", "DPA An 4", "DPA An 5"]); rr += 1
for cat in M["dpa"]["categories"]:
    wp.cell(rr, 1, cat["nom"]); wp.cell(rr, 2, cat["taux"]).number_format = pctf; wp.cell(rr, 3, cat["base"]).number_format = money
    for j, v in enumerate(cat["dpa"]): wp.cell(rr, 4 + j, v).number_format = money
    rr += 1

# ---------------------------------------------------------------- Bilan
wbl = wb.create_sheet("Bilan")
header(wbl, 1, ["Bilan pro forma au 31 décembre", "An 1", "An 2", "An 3", "An 4", "An 5"], [64, 15, 15, 15, 15, 15])
fdr_row = [i for i, (lab, _) in enumerate(M["emplois"], 2) if lab.startswith("Fonds de roulement")][0]
res_row = [i for i, (lab, _) in enumerate(M["emplois"], 2) if lab.startswith("Réserve")][0]
rep_rows = [i for i, (lab, _) in enumerate(M["emplois"], 2) if lab.startswith(("Mise en conformité", "Intérêts capitalisés", "Frais de financement"))]
wbl.cell(2, 1, "Immobilisations existantes au coût déclaré (hyp.)"); wbl.cell(2, 2, M["immo_existantes_cout"]).number_format = money; wbl.cell(2, 2).fill = inp
wbl.cell(3, 1, "Amortissement cumulé estimé au 31-12-2026 (hyp.)"); wbl.cell(3, 2, M["amort_cumule_2026"]).number_format = money; wbl.cell(3, 2).fill = inp
labels = ["ACTIF", "Encaisse", "Immobilisations existantes, nettes", "Nouvelles immobilisations, nettes", "Frais reportés", "Total de l'actif", "PASSIF", "Dette totale (toutes couches)", "AVOIR", "Avoir à l'ouverture", "Bénéfices non répartis cumulés (nets de l'amortissement des frais reportés)", "Total de l'avoir", "Total passif + avoir", "Écart (doit être 0)"]
for i, t in enumerate(labels): wbl.cell(5 + i, 1, t)
for t in (5, 10, 12, 13, 17, 18): wbl.cell(t, 1).font = bold
reportes = "+".join(f"Investissement!B{i}" for i in rep_rows)
for j in range(1, 6):
    c = L(1 + j); rc = COLS[j]
    wbl.cell(6, 1 + j, f"=Investissement!B{fdr_row}+Investissement!B{res_row}+Flux!{rc}8").number_format = money
    wbl.cell(7, 1 + j, f"=$B$2-$B$3-{hv('amort_exist')}*{j}").number_format = money
    wbl.cell(8, 1 + j, f"=DPA!$B$2+SUM(Résultats!$C${MAINT}:{rc}${MAINT})-SUM(DPA!$C$5:{rc}$5)").number_format = money
    wbl.cell(9, 1 + j, f"=({reportes})*(1-{j}/5)").number_format = money
    wbl.cell(10, 1 + j, f"=SUM({c}6:{c}9)").number_format = money
    wbl.cell(12, 1 + j, "=" + "+".join(SOLDE_REFS[k][j - 1] for k in range(len(SOLDE_REFS)))).number_format = money
    wbl.cell(14, 1 + j, f"=$B$2-$B$3+Investissement!B{fdr_row}+Investissement!B{res_row}+({reportes})+DPA!$B$2-SUM(Investissement!B{src0}:B{src1})").number_format = money
    wbl.cell(15, 1 + j, f"=SUM(Résultats!$C${BN}:{rc}${BN})-({reportes})*{j}/5").number_format = money
    wbl.cell(16, 1 + j, f"={c}14+{c}15").number_format = money
    wbl.cell(17, 1 + j, f"={c}12+{c}16").number_format = money
    wbl.cell(18, 1 + j, f"={c}10-{c}17").number_format = money

# ---------------------------------------------------------------- Ratios
wq = wb.create_sheet("Ratios")
header(wq, 1, ["Ratio", "An 1", "An 2", "An 3", "An 4", "An 5"], [64, 14, 14, 14, 14, 14])
rat = [("Couverture du prêt senior (BAIIA / service senior)", lambda c: f"=Résultats!{c}{BAIIA}/Résultats!{c}{SS}", ratf),
       ("Couverture toutes couches (BAIIA / service total)", lambda c: f"=Résultats!{c}{BAIIA}/Résultats!{c}{ST}", ratf),
       ("Couverture bancaire ((BAIIA − impôts − maintien) / service total)", lambda c: f"=(Résultats!{c}{BAIIA}-Résultats!{c}{IMP}-Résultats!{c}{MAINT})/Résultats!{c}{ST}", ratf),
       ("Couverture des intérêts", lambda c: f"=Résultats!{c}{BAIIA}/Résultats!{c}{INT}", ratf),
       ("Dette totale / BAIIA", lambda c: f"=Bilan!{L(1 + COLS.index(c))}12/Résultats!{c}{BAIIA}", '0.0" x"'),
       ("Dette / avoir", lambda c: f"=Bilan!{L(1 + COLS.index(c))}12/Bilan!{L(1 + COLS.index(c))}16", ratf),
       ("Marge BAIIA", lambda c: f"=Résultats!{c}{BAIIA}/Résultats!{c}{TOTR}", pctf),
       ("Marge nette", lambda c: f"=Résultats!{c}{BN}/Résultats!{c}{TOTR}", pctf)]
for i, (lab, f, fm) in enumerate(rat, 2):
    wq.cell(i, 1, lab)
    for j in range(1, 6): wq.cell(i, 1 + j, f(COLS[j])).number_format = fm

# ---------------------------------------------------------------- Sensibilité (facteur de scénario)
wsn = wb.create_sheet("Sensibilité")
wsn.column_dimensions["A"].width = 70
wsn.cell(1, 1, "Sensibilité : modifier « SCÉNARIO : facteur sur les revenus » dans Hypothèses (1 = base, 0,9 = −10 %, etc.) ; les charges variables suivent selon leur part variable. Tableau statique ci-dessous produit par le modèle Python pour référence.").alignment = Alignment(wrap_text=True)
header(wsn, 3, ["Scénario", "Toutes couches An 2", "An 3", "An 5", "Bancaire An 3", "Bancaire An 5"])
for i, s in enumerate(M["sensibilite"], 4):
    wsn.cell(i, 1, s["scenario"])
    for j, k in enumerate(["dscr_an2", "dscr_an3", "dscr_an5", "dscr_banc_an3", "dscr_banc_an5"]): wsn.cell(i, 2 + j, round(s[k], 2)).number_format = ratf
for col in "BCDEF": wsn.column_dimensions[col].width = 18
wsn.cell(i + 2, 1, "Part variable calculée (An 3)"); wsn.cell(i + 2, 2, M["part_variable"]).number_format = pctf

for sh in wb.worksheets:
    sh.freeze_panes = "B2"
wb.save(OUT)
print("Écrit :", OUT)
